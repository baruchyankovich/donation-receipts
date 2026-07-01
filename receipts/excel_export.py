"""Export receipts to an Excel workbook (falls back to CSV without openpyxl)."""

from __future__ import annotations

from pathlib import Path

from .formatting import parse_amount
from .models import Receipt

_HEADERS = (
    "מספר קבלה",
    "תאריך",
    "שם התורם",
    "כתובת",
    "טלפון",
    "אימייל",
    "עבור",
    "פרט מוצר",
    "שווי (₪)",
    "שווי במילים",
    "הערות",
)
_COLUMN_WIDTHS = (11, 12, 22, 24, 14, 20, 16, 26, 10, 26, 20)


def _rows(receipts: list[Receipt]) -> list[list[object]]:
    return [[getattr(r, col) for col in Receipt.COLUMNS] for r in receipts]


def export(receipts: list[Receipt], path: str | Path) -> Path:
    """Write ``receipts`` to ``path`` as .xlsx, or .csv if openpyxl is missing."""
    path = Path(path)
    try:
        return _export_xlsx(receipts, path)
    except ImportError:
        return _export_csv(receipts, path.with_suffix(".csv"))


def _export_xlsx(receipts: list[Receipt], path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "קבלות"
    sheet.sheet_view.rightToLeft = True

    sheet.append(list(_HEADERS))
    header_fill = PatternFill("solid", fgColor="000000")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    for row in _rows(receipts):
        sheet.append(row)
        total += parse_amount(row[Receipt.COLUMNS.index("value_num")])

    sheet.append([])
    total_row = [""] * len(_HEADERS)
    total_row[6] = "סה״כ:"
    total_row[8] = round(total, 2)
    sheet.append(total_row)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    workbook.save(str(path))
    return path


def _export_csv(receipts: list[Receipt], path: Path) -> Path:
    import csv

    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(_HEADERS)
        writer.writerows(_rows(receipts))
    return path
